import openpyxl
from datetime import datetime

EXCEL_FILE_PATH = "C:\\Users\\siddh\\OneDrive\\Desktop\\PYTHON PROG\\VITYARTHI-PROJECT-SEM-1\\Vityarthi-Project-1st-Sem\\attendance (3).xlsx"

def load_attendance_workbook(path):
    
    #Load the Excel workbook; if it doesn't exist, create a new one with headers.
    
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Set up headers: ID, Name (no date columns yet)
        sheet.append(["ID", "Name"])
    return workbook, sheet

def save_attendance_workbook(workbook, path):
    
    #Save the workbook to the specified path.
    
    workbook.save(path)
    print(f"Attendance data saved to {path}.")

def add_student(sheet):
    
    #Prompt for a new student name and ID, then append to the sheet.
    #Prevent duplicate IDs.
    
    name = input("Enter student name: ")
    # Loop until a valid integer ID is entered
    while True:
        id_input = input("Enter student ID (integer): ")
        try:
            student_id = int(id_input)
            break
        except ValueError:
            print("Invalid ID. Please enter an integer.")
    # Check for duplicate ID
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if student_id == row[0]:
            print("ID already exists. Student not added.")
            return
    # Determine number of existing date columns (if any)
    max_date_cols = sheet.max_column - 2  # subtract ID and Name columns
    # Append new student row with blank entries for existing dates
    sheet.append([student_id, name] + [""] * max_date_cols)
    print(f"Student {name} (ID: {student_id}) added.")

def find_date_column(sheet, date_str):
    
    #Find the column index for the given date header, or return None if not found.
    #Date columns start from column 3 onward.
    
    for col in range(3, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date_str:
            return col
    return None

def add_date_column(sheet, date_str):
    
  #Add a new date column with header `date_str`, and fill existing rows with blank.
    
    col = sheet.max_column + 1
    sheet.cell(row=1, column=col).value = date_str
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=col).value = ""
    return col

def record_attendance(sheet):
    
    #Prompt for a date, student ID, and status (Present/Absent), then record it.
    
    date_input = input("Enter date (YYYY-MM-DD): ")
    try:
        date_obj = datetime.strptime(date_input, "%Y-%m-%d").date()
    except ValueError:
        print("Invalid date format. Use YYYY-MM-DD.")
        return
    date_str = date_obj.isoformat()
    col = find_date_column(sheet, date_str)
    if col is None:
        col = add_date_column(sheet, date_str)
        print(f"Added new date column: {date_str}")
    try:
        student_id = int(input("Enter student ID: "))
    except ValueError:
        print("Invalid ID. Please enter an integer.")
        return
    status = input("Enter status (Present/Absent): ").strip().capitalize()
    if status not in ("Present", "Absent"):
        print("Invalid status. Enter 'Present' or 'Absent'.")
        return
    # Locate student row and record status
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == student_id:
            sheet.cell(row=row, column=col).value = status
            print(f"Recorded {status} for student ID {student_id} on {date_str}.")
            return
    print("Student ID not found.")

def show_attendance(sheet):
    
    #Display attendance. If an ID is given, show only that row; otherwise show all.
    
    id_input = input("Enter student ID to view attendance (or leave blank for all): ")
    if id_input == "":
        # Print all rows (including header)
        for row in sheet.iter_rows(values_only=True):
            print(row)
    else:
        try:
            student_id = int(id_input)
        except ValueError:
            print("Invalid ID.")
            return
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id:
                print(row)
                return
        print("Student ID not found.")

def main():
    workbook, sheet = load_attendance_workbook(EXCEL_FILE_PATH)
    while True:
        print("\nOptions: 1) Add Student  2) Record Attendance  3) Show Attendance  4) Save and Exit")
        choice = input("Enter choice (1-4): ")
        if choice == "1":
            add_student(sheet)
        elif choice == "2":
            record_attendance(sheet)
        elif choice == "3":
            show_attendance(sheet)
        elif choice == "4":
            save_attendance_workbook(workbook, EXCEL_FILE_PATH)
            break
        else:
            print("Invalid choice. Try again.")

if __name__ == "__main__":
    main()
